# noinspection SpellCheckingInspection
__author__ = 'wookjae.jo'

import json
import logging
import os
import pickle
from datetime import date
from typing import *

import jsons
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from indexes import InterestIndexes
from .backtest import AbcBacktest, BacktestEvent, BuyEvent, SellEvent, DailyLog
from .common import get_fl_map, get_name


def date_to_str(d: date):
    return d.strftime('%Y-%m-%d')


class XlsxExporter:

    def __init__(self, backtest: AbcBacktest, target_path):
        self.backtest = backtest
        self.workbook = openpyxl.Workbook()
        self.target_path = target_path

    @classmethod
    def load_from(cls, pickle_or_json: str):
        if pickle_or_json.endswith('.pickle'):
            with open(pickle_or_json, 'rb') as f:
                backteset = pickle.load(f)
        elif pickle_or_json.endswith('.json'):
            with open(pickle_or_json, 'r', encoding='utf-8') as f:
                backteset = pickle.load(f)
        else:
            raise RuntimeError(f'Not supported file extension: {pickle_or_json}')

        target_dir = os.path.abspath(os.path.join(os.path.abspath(pickle_or_json), os.path.pardir))
        return XlsxExporter(
            backtest=backteset,
            target_path=os.path.join(target_dir, f'Result-{os.path.basename(target_dir)}.xlsx')
        )

    def create_table_sheet(self, title, headers: List, rows: List[List], index=None):
        sheet = self.workbook.create_sheet(title=title, index=index)
        font_header = Font(name='맑은 고딕', bold=True, color='ffffad')
        align_header = Alignment(horizontal='center')
        fill_header = PatternFill('solid', '4a4a4a')

        row_pos = 1
        for i in range(len(headers)):
            cell = sheet.cell(row_pos, i + 1, headers[i])
            cell.font = font_header
            cell.alignment = align_header
            cell.fill = fill_header

        row_pos += 1
        font_row = Font(name='맑은 고딕')
        for i in range(len(rows)):
            for j in range(len(rows[i])):
                cell = sheet.cell(row_pos + i, j + 1, rows[i][j])
                cell.font = font_row

        column_widths = []
        for row in sheet.rows:
            for i, cell in enumerate(row):
                width = 0
                for c in str(cell.value):
                    if ord(c) > 255:
                        width += 2.5
                    else:
                        width += 1.5

                if len(column_widths) > i:
                    if width > column_widths[i]:
                        column_widths[i] = width
                else:
                    column_widths.append(width)

        for i, column_width in enumerate(column_widths):
            sheet.column_dimensions[get_column_letter(i + 1)].width = column_width

    def export(self):
        logging.info('Exporting report...')
        for sheet_name in self.workbook.sheetnames:
            self.workbook.remove(self.workbook[sheet_name])

        def total_eval(_daily_log: DailyLog):
            return _daily_log.deposit + _daily_log.holding_eval

        logging.info('Making ranking sheet...')
        events_by_code: [str, List[BacktestEvent]] = {}
        for event in self.backtest.events:
            if event.code not in events_by_code:
                events_by_code.update({event.code: []})

            events_by_code.get(event.code).append(event)

        revenues_by_code = {}
        revenue_rates_by_code = {}
        for code in events_by_code:
            events: List[BacktestEvent] = events_by_code.get(code)
            revenue = 0
            seed = 0
            for event in events:
                order_total = event.price * event.quantity
                if isinstance(event, BuyEvent):
                    revenue -= order_total
                    seed += order_total
                elif isinstance(event, SellEvent):
                    revenue += order_total
                else:
                    RuntimeError('WTF')

            if seed:
                revenues_by_code.update({code: revenue})
                revenue_rates_by_code.update({code: revenue / seed * 100})

        rows = []
        fl_map = get_fl_map(list(revenues_by_code.keys()), self.backtest.begin, self.backtest.end)
        for code in revenues_by_code:
            first, last = fl_map.get(code)
            revenue = round(revenues_by_code.get(code), 2)
            revenue_rate = round(revenue_rates_by_code.get(code), 2)
            margin_rate = round((last - first) / first * 100, 2)
            row = [
                code, get_name(code),
                revenue,
                revenue_rate,
                first,
                last,
                margin_rate,
                revenue_rate - margin_rate
            ]
            rows.append(row)

        # 수익금으로 정렬
        rows.sort(key=lambda x: x[-1], reverse=True)
        self.create_table_sheet(
            'ranking',
            headers=['종목코드', '종목명', '수익금', '수익율(%)', '시작가', '종료가', '변동율(%)', '수익율-변동율(%)'],
            rows=rows, index=2
        )

        logging.info('Making events sheet...')
        headers = [
            '일시', '구분', '종목코드', '종목명', '가격',
            '수량', '주문총액', '비고',
            '수익율(%)'
        ]
        rows = [[
            evt.when, evt.__class__.__name__, evt.code, get_name(evt.code), evt.price,
            evt.quantity, evt.price * evt.quantity, evt.comment,
            round(evt.revenue_percent, 2)
            if isinstance(evt, SellEvent) else ''
        ] for evt in self.backtest.events]
        self.create_table_sheet('events', headers=headers, rows=rows, index=2)

        logging.info('Making daily sheet...')
        headers = [
            '날짜', '예수금', '보유 종목 평가금액', '총 평가금액',
            '전일대비(%)', '비고',
            '코스피', '코스피 전일대비(%)',
            '코스닥', '코스피 전일대비(%)',
            'KRX 300', 'KRX 300(%)'
        ]

        rows = []
        indexes = InterestIndexes.load(fromdate=self.backtest.begin, todate=self.backtest.end)
        for i in range(len(self.backtest.daily_logs)):
            dl = self.backtest.daily_logs[i]
            if i > 0:
                dl_prv = self.backtest.daily_logs[i - 1]
                margin = (total_eval(dl) - total_eval(dl_prv)) / total_eval(dl_prv) * 100
            else:
                margin = (total_eval(dl) - self.backtest.initial_deposit) / self.backtest.initial_deposit * 100

            try:
                # noinspection PyUnboundLocalVariable
                kospi_prev, kosdaq_prev, krx_300_prev = kospi, kosdaq, krx_300_prev
            except:
                kospi_prev, kosdaq_prev, krx_300_prev = 0, 0, 0

            kospi = indexes.kospi.data.get(dl.date).close
            kosdaq = indexes.kosdaq.data.get(dl.date).close
            krx_300 = indexes.krx_300.data.get(dl.date).close

            if not kospi_prev:
                kospi_prev = kospi

            if not kosdaq_prev:
                kosdaq_prev = kosdaq

            if not krx_300_prev:
                krx_300_prev = krx_300

            try:
                kospi_margin_rate = round((kospi - kospi_prev) / kospi_prev * 100, 2)
                kosdaq_margin_rate = round((kosdaq - kosdaq_prev) / kosdaq_prev * 100, 2)
                krx_300_margin_rate = round((krx_300 - krx_300_prev) / krx_300_prev * 100, 1)
            except:
                kospi_margin_rate = 0
                kosdaq_margin_rate = 0
                krx_300_margin_rate = 0

            rows.append([
                dl.date, round(dl.deposit), round(dl.holding_eval), round(total_eval(dl)),
                round(margin, 2), dl.comment,
                kospi, kospi_margin_rate,
                kosdaq, kosdaq_margin_rate,
                krx_300, krx_300_margin_rate
            ])

        self.create_table_sheet('daily', headers=headers, rows=rows, index=1)

        logging.info('Making summary sheet...')
        headers = [
            '시작일', '종료일', '구동시간(sec)', '최초 예수금',
            '수익금',
            '코스피 증감율(%)', '코스닥 증감율(%)', 'KRX 300 증감율(%)'
        ]

        margin_percentage_list = []
        for code in fl_map:
            first, last = fl_map.get(code)
            margin_percentage_list.append(((last - first) / first) * 100)

        row = [
            self.backtest.begin, self.backtest.end, (self.backtest.finish_time - self.backtest.start_time).seconds,
            self.backtest.initial_deposit,
            round(self.backtest.account.deposit - self.backtest.initial_deposit),
            # round((self.backtest.account.deposit - self.backtest.initial_deposit) / self.backtest.initial_deposit * 100, 2),
            # round(sum(margin_percentage_list) / len(margin_percentage_list), 2),
        ]

        indexes = InterestIndexes.load(fromdate=self.backtest.begin, todate=self.backtest.end)
        kospi_f = list(indexes.kospi.data.values())[0].close
        kospi_l = list(indexes.kospi.data.values())[-1].close
        kospi_margin_rate = (kospi_l - kospi_f) / kospi_f * 100

        kosdaq_f = list(indexes.kosdaq.data.values())[0].close
        kosdaq_l = list(indexes.kosdaq.data.values())[-1].close
        kosdaq_margin_rate = (kosdaq_l - kosdaq_f) / kospi_f * 100

        krx_300_f = list(indexes.krx_300.data.values())[0].close
        krx_300_l = list(indexes.krx_300.data.values())[-1].close
        krx_300_margin_rate = (krx_300_l - krx_300_f) / krx_300_f * 100

        row += [
            round(kospi_margin_rate, 2), round(kosdaq_margin_rate, 2), round(krx_300_margin_rate, 2)
        ]

        rows = [row]
        self.create_table_sheet('summary', headers=headers, rows=rows, index=0)
        logging.info(f'Saving workbook in {self.target_path}')
        self.workbook.save(self.target_path)

        with open(os.path.join(os.path.dirname(self.target_path), '.json'), 'w', encoding='utf-8') as f:
            json.dump(jsons.dumps(self.backtest), f, indent=2)

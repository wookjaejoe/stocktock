from __future__ import annotations

import csv
import json
import logging
import os
import pickle
from dataclasses import dataclass
from datetime import date, timedelta, datetime, time
from enum import Enum
from typing import *

import jsons

import database.charts
import database.metrics
import database.stocks
from common.metric import MaCalculator
from strategy.strategies import Over5MaStrategy

with database.stocks.StockTable() as stock_table:
    stocks = [stock for stock in stock_table.all() if '스팩' not in stock.name]


def date_to_str(d: date):
    return d.strftime('%Y-%m-%d')


def get_name(code: str):
    for stock in stocks:
        if stock.code == code:
            return stock.name


class BacktestEventType(Enum):
    BUY = 0
    SELL = 1


@dataclass
class BacktestEvent:
    when: datetime
    code: str
    type: BacktestEventType
    price: int
    count: int
    description: str


@dataclass
class BacktestSellEvent(BacktestEvent):
    bought_event: BacktestEvent


@dataclass
class Holding:
    bought_event: BacktestEvent


@dataclass
class DailyLog:
    date: date
    deposit: int  # 예수금
    holding_eval: int  # 보유 종목 평가금액
    holding_count: int  # 보유 종목 개수
    description: str  # 추가 정보


@dataclass
class BacktestReport:
    begin: date  # 시작일
    end: date  # 종료일
    running_time: int  # 구동 시간(초)
    earn_line: float  # 익절라인
    stop_line: float  # 손절라인


class Cache:
    pass


cache = Cache()


class DayCandleFetcher:
    def fetch(self):
        pass


class MinuteCandleFetcher:
    def fetch(self):
        pass


# noinspection PyMethodMayBeStatic
class Backtest:

    def __init__(
            self,
            begin: date,
            end: date,

            initial_deposit: int,

            limit_holding_count: int,
            limit_buy_amount: int,
            limit_holding_days: int,

            earn_line: float,
            stop_line: float,

            tax_percent: float,
            fee_percent: float,
    ):
        self.strategy = Over5MaStrategy(earn_line, stop_line)
        self.begin = begin
        self.end = end

        self.limit_holding_count = limit_holding_count
        self.limit_buy_amount = limit_buy_amount
        self.limit_holding_days = limit_holding_days

        self.earn_line = earn_line
        self.stop_line = stop_line

        self.tax_percent = tax_percent
        self.fee_percent = fee_percent

        self.initial_deposit = initial_deposit
        self.deposit = initial_deposit
        self.events: List[BacktestEvent] = []
        self.holdings: Dict[str, Holding] = {}
        self.daily_logs: List[DailyLog] = []

        self.start_time: Optional[datetime] = None
        self.finish_time: Optional[datetime] = None

    def try_buy(self, when: datetime, code: str, price: int, description: str):
        assert code not in self.holdings

        count = int(self.limit_buy_amount / price)
        if price * count > self.deposit:
            return

        event = BacktestEvent(
            type=BacktestEventType.BUY,
            when=when, code=code, price=price, count=count,
            description=description
        )
        self.events.append(event)
        self.holdings.update({code: Holding(bought_event=event)})
        self.deposit -= price * count
        self.deposit -= int(price * count * self.fee_percent / 100)  # 수수료 차감

    def try_sell(self, code: str, when: datetime, price: int, description: str):
        assert code in self.holdings
        count = self.holdings.get(code).bought_event.count
        event = BacktestSellEvent(
            type=BacktestEventType.SELL,
            when=when, code=code, price=price, count=count,
            description=description,
            bought_event=self.holdings.get(code).bought_event
        )
        self.events.append(event)
        del self.holdings[code]
        self.deposit += price * count
        self.deposit -= int(price * count * self.fee_percent / 100)  # 수수료 차감
        self.deposit -= int(price * count * self.tax_percent / 100)  # 매도세 차감

    def run(self, d: date):
        logging.info(f'The day is {date_to_str(d)}')

        day_candles_map = {}
        with database.charts.DayCandlesTable() as day_candles_table:
            if not day_candles_table.exists(date=d):
                logging.info(f'Not a business day.')
                return

            logging.info('Fetching day candles...')
            for day_candle in day_candles_table.find_all(
                    codes=[stock.code for stock in stocks],
                    begin=d - timedelta(days=200),
                    end=d
            ):
                if day_candle.code not in day_candles_map:
                    day_candles_map.update({day_candle.code: []})

                day_candles_map.get(day_candle.code).append(day_candle)

        logging.info('Making whitelist...')
        whitelist = []
        for stock in [stock for stock in stocks if stock.code in day_candles_map]:
            day_candles = day_candles_map.get(stock.code)
            ma_calc = MaCalculator([candle.close for candle in day_candles])

            try:
                ma_5_yst = ma_calc.get(5, pos=-1)
                ma_20_yst = ma_calc.get(20, pos=-1)
                ma_20_yst_2 = ma_calc.get(20, pos=-2)
                ma_60_yst = ma_calc.get(60, pos=-1)
                ma_120_yst = ma_calc.get(120, pos=-1)
            except:
                continue

            # 배열상태 and 20MA 기울기 (+) and # !연속음봉
            if ma_20_yst > ma_5_yst \
                    and ma_20_yst > ma_60_yst > ma_120_yst \
                    and ma_20_yst > ma_20_yst_2 \
                    and len([candle for candle in day_candles[-5:] if candle.open < candle.close]) > 0:
                whitelist.append(stock.code)

        logging.info(f'Streaming minute candles...')
        with database.charts.MinuteCandlesTable(d) as minute_candles_table:
            minute_candles = minute_candles_table.find_all(codes=whitelist, use_yield=True)

        logging.info(f'Looking details at tbe minute candles for {len(whitelist)} stocks...')
        for minute_candle in minute_candles:
            code = minute_candle.code
            cur_price = minute_candle.close
            when = datetime.combine(minute_candle.date, minute_candle.time)

            if code in self.holdings.keys():  # 보유중
                buy_price = self.holdings.get(code).bought_event.price
                margin_percent = (cur_price - buy_price) / buy_price * 100
                day_candles_after_bought = [dc for dc in day_candles_map.get(code) if
                                            dc.date >= self.holdings.get(code).bought_event.when.date()]
                holding_days = len(day_candles_after_bought)
                # 손익절 라인 확인
                if margin_percent >= self.earn_line:
                    self.try_sell(code=code, when=when, price=cur_price, description='익절')
                elif margin_percent <= self.stop_line:
                    self.try_sell(code=code, when=when, price=cur_price, description='손절')
                # 매도 5+ 거래일 지난 15:00 이후
                elif holding_days >= 5 and when.time() > time(14, 30):
                    vol_avg = sum([dc.vol for dc in day_candles_after_bought]) / len(day_candles_after_bought)
                    # 변동성 1% 미만이면, 매도
                    if abs(margin_percent) < 1:
                        self.try_sell(code=code, when=when, price=cur_price, description='미변동')
                    # 평균 거래량이 매수 당시 거래량보다 작으면 매도 todo: 확인
                    # elif day_candles_after_bought[0].vol > vol_avg:
                    #     self.try_sell(code=code, when=when, price=cur_price, description='매수당시거래량 > 평균거래량')
                    # 보유 기간 제한 20일
                    elif holding_days >= self.limit_holding_days:
                        self.try_sell(code=code, when=when, price=cur_price,
                                      description=f'보유 기간 제한({self.limit_holding_days})일 초과')

                self.strategy.check_and_sell()
            else:  # 미보유
                self.strategy.check_and_buy(
                    day_candles=day_candles_map.get(minute_candle.code),
                    cur_price=minute_candle.close,
                    buy=lambda: self.try_buy(
                        code=code,
                        when=when,
                        price=cur_price,
                        description='매수 조건 만족'
                    )
                )

    def evaluate(self, d: date):
        day_candles_map = {}
        with database.charts.DayCandlesTable() as day_candles_table:
            day_candles = day_candles_table.find_all(codes=list(self.holdings.keys()), begin=d, end=d)

        for day_candle in day_candles:
            if day_candle.code not in day_candles_map:
                day_candles_map.update({day_candle.code: []})

            day_candles_map.get(day_candle.code).append(day_candle)

        holding_eval = 0
        no_candle_codes = []
        for code in self.holdings:
            bought_event = self.holdings.get(code).bought_event
            if code in day_candles_map:
                # 종가 * 보유 개수
                holding_eval += day_candles_map.get(code)[-1].close * bought_event.count
            else:
                holding_eval += bought_event.price * bought_event.count
                no_candle_codes.append(code)

        holding_eval -= holding_eval * self.tax_percent / 100  # 매도세 반영

        return DailyLog(
            date=d,
            deposit=self.deposit,
            holding_count=len(self.holdings),
            holding_eval=holding_eval,
            description=f'{len(no_candle_codes)} stocks have no day candle.'
        )

    def start(self):
        self.start_time = datetime.now()
        for d in [self.begin + timedelta(days=i) for i in range((self.end - self.begin).days + 1)]:
            self.run(d)
            self.daily_logs.append(self.evaluate(d))

        self.finish_time = datetime.now()

        target_dir = os.path.join('reports', datetime.now().strftime('%Y%m%d_%H%M%S'))
        os.makedirs(target_dir, exist_ok=True)
        logging.info('Dumping result...')
        self.dump(target_dir)
        logging.info('Saving result...')
        self.save_report(os.path.join(target_dir, f'Result-{os.path.basename(target_dir)}.xlsx'))
        logging.info(f'FINISHED: took {(self.finish_time - self.start_time).seconds} seconds.')

    def dump(self, target_dir) -> str:
        logging.info('Dumping result...')
        json_path = os.path.join(target_dir, 'backtest.json')
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(jsons.dump(self), f)

        pickle_path = os.path.join(target_dir, 'backtest.pickle')
        with open(pickle_path, 'wb') as f:
            pickle.dump(self, f)

        return target_dir

    def save_report(self, target_path):
        XlsxExporter(backtest=self, target_path=target_path).export()


import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from indexes import InterestIndexes


class XlsxExporter:

    def __init__(self, backtest: Backtest, target_path):
        self.backteset = backtest
        self.workbook = openpyxl.Workbook()
        self.target_path = target_path

    @classmethod
    def load_from(cls, pickle_or_json: str):
        if pickle_or_json.endswith('.pickle'):
            with open(pickle_or_json, 'rb') as f:
                backteset = pickle.load(f)
        elif pickle_or_json.endswith('.json'):
            with open(pickle_or_json, 'r', encoding='utf-8') as f:
                backteset = pickle.load(f)
        else:
            raise RuntimeError(f'Not supported file extension: {pickle_or_json}')

        target_dir = os.path.abspath(os.path.join(os.path.abspath(pickle_or_json), os.path.pardir))
        return XlsxExporter(
            backtest=backteset,
            target_path=os.path.join(target_dir, f'Result-{os.path.basename(target_dir)}.xlsx')
        )

    def create_table_sheet(self, title, headers: List, rows: List[List], index=None):
        sheet = self.workbook.create_sheet(title=title, index=index)
        font_header = Font(name='맑은 고딕', bold=True, color='ffffad')
        align_header = Alignment(horizontal='center')
        fill_header = PatternFill('solid', '4a4a4a')

        row_pos = 1
        for i in range(len(headers)):
            cell = sheet.cell(row_pos, i + 1, headers[i])
            cell.font = font_header
            cell.alignment = align_header
            cell.fill = fill_header

        row_pos += 1
        font_row = Font(name='맑은 고딕')
        for i in range(len(rows)):
            for j in range(len(rows[i])):
                cell = sheet.cell(row_pos + i, j + 1, rows[i][j])
                cell.font = font_row

        column_widths = []
        for row in sheet.rows:
            for i, cell in enumerate(row):
                width = 0
                for c in str(cell.value):
                    if ord(c) > 255:
                        width += 2.5
                    else:
                        width += 1.5

                if len(column_widths) > i:
                    if width > column_widths[i]:
                        column_widths[i] = width
                else:
                    column_widths.append(width)

        for i, column_width in enumerate(column_widths):
            sheet.column_dimensions[get_column_letter(i + 1)].width = column_width

    def export(self):
        for sheet_name in self.workbook.sheetnames:
            self.workbook.remove(self.workbook[sheet_name])

        indexes = InterestIndexes.load(fromdate=self.backteset.begin, todate=self.backteset.end)

        def total_eval(_daily_log: DailyLog):
            return _daily_log.deposit + _daily_log.holding_eval

        ########## events ##########
        headers = [
            '일시', '구분', '종목코드', '종목명', '가격',
            '수량', '주문총액', '비고',
            '수익율(%)'
        ]
        rows = [[
            evt.when, evt.type.name, evt.code, get_name(evt.code), evt.price,
            evt.count, evt.price * evt.count, evt.description,
            round((evt.price - evt.bought_event.price) / evt.bought_event.price * 100, 2)
            if isinstance(evt, BacktestSellEvent) else ''
        ] for evt in self.backteset.events]
        self.create_table_sheet('events', headers=headers, rows=rows, index=2)

        ########## daily ##########
        headers = [
            '날짜', '예수금', '보유 종목 평가금액', '총 평가금액',
            '전일대비(%)', '비고',
            '코스피', '코스피 전일대비(%)',
            '코스닥', '코스피 전일대비(%)',
            'KRX 300', 'KRX 300(%)'
        ]

        rows = []
        for i in range(len(self.backteset.daily_logs)):
            dl = self.backteset.daily_logs[i]
            if i > 0:
                dl_prv = self.backteset.daily_logs[i - 1]
                margin = (total_eval(dl) - total_eval(dl_prv)) / total_eval(dl_prv) * 100
            else:
                margin = (total_eval(dl) - self.backteset.initial_deposit) / self.backteset.initial_deposit * 100

            try:
                kospi = indexes.kospi.data.get(dl.date).close
                kosdaq = indexes.kosdaq.data.get(dl.date).close
                krx_300 = indexes.krx_300.data.get(dl.date).close

                kospi_bf = indexes.kospi.data.get(dl.date - timedelta(days=1)).close
                kosdaq_bf = indexes.kosdaq.data.get(dl.date - timedelta(days=1)).close
                krx_300_bf = indexes.krx_300.data.get(dl.date - timedelta(days=1)).close

                rows.append([
                    dl.date, round(dl.deposit), round(dl.holding_eval), round(total_eval(dl)),
                    round(margin, 2), dl.description,
                    kospi, round((kospi - kospi_bf) / kospi_bf * 100, 2),
                    kosdaq, round((kosdaq - kosdaq_bf) / kosdaq_bf * 100, 2),
                    krx_300, round((krx_300 - krx_300_bf) / krx_300_bf * 100, 1)
                ])
            except:
                pass

        self.create_table_sheet('daily', headers=headers, rows=rows, index=1)

        ########## summary ##########
        headers = [
            '시작일', '종료일', '구동시간(sec)', '최초 예수금',
            '익절라인(%)', '손절라인(%)', '매매 수수료(%)', '매도 세금(%)',
            '수익금',
            '수익율(%)',
            '코스피 증감율(%)', '코스닥 증감율(%)', 'KRX 300 증감율(%)'
        ]

        kospi_f = list(indexes.kospi.data.values())[0].close
        kospi_l = list(indexes.kospi.data.values())[-1].close
        kospi_margin = (kospi_l - kospi_f) / kospi_f

        kosdaq_f = list(indexes.kosdaq.data.values())[0].close
        kosdaq_l = list(indexes.kosdaq.data.values())[-1].close
        kosdaq_margin = (kosdaq_l - kosdaq_f) / kospi_f

        krx_300_f = list(indexes.krx_300.data.values())[0].close
        krx_300_l = list(indexes.krx_300.data.values())[-1].close
        krx_300_margin = (krx_300_l - krx_300_f) / krx_300_f

        final_eval = total_eval(self.backteset.daily_logs[-1])
        rows = [[
            self.backteset.begin, self.backteset.end, (self.backteset.finish_time - self.backteset.start_time).seconds,
            self.backteset.initial_deposit,
            self.backteset.earn_line, self.backteset.stop_line, self.backteset.fee_percent, self.backteset.tax_percent,
            round(final_eval - self.backteset.initial_deposit),
            round((final_eval - self.backteset.initial_deposit) / self.backteset.initial_deposit * 100, 2),
            round(kospi_margin, 2), round(kosdaq_margin, 2), round(krx_300_margin, 2)
        ]]
        self.create_table_sheet('summary', headers=headers, rows=rows, index=0)
        logging.info(f'Saving workbook in {self.target_path}')
        self.workbook.save(self.target_path)


def write_csv(path: str, headers: List[Any], values: List[List[Any]]):
    with open(path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f, quoting=csv.QUOTE_NONNUMERIC)
        writer.writerow(headers)
        writer.writerows(values)

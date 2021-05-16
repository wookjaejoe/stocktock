import csv
import os
from typing import *

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


class TableSheetGenerator:
    def __init__(self, workbook: openpyxl.Workbook, title: str):
        self.workbook = workbook
        self.sheet = self.workbook.create_sheet(title=title)

    def generate(self, headers: List, rows: List[List]):
        font_header = Font(name='맑은 고딕', bold=True, color='ffffad')
        align_header = Alignment(horizontal='center')
        fill_header = PatternFill('solid', '4a4a4a')

        row_pos = 1
        for i in range(len(headers)):
            cell = self.sheet.cell(row_pos, i + 1, headers[i])
            cell.font = font_header
            cell.alignment = align_header
            cell.fill = fill_header

        row_pos += 1
        font_row = Font(name='맑은 고딕')
        for i in range(len(rows)):
            for j in range(len(rows[i])):
                cell = self.sheet.cell(row_pos + i, j + 1, rows[i][j])
                cell.font = font_row

        column_widths = []
        for row in self.sheet.rows:
            for i, cell in enumerate(row):
                width = 0
                for c in cell.value:
                    if ord(c) > 255:
                        width += 2.5
                    else:
                        width += 1.5

                if len(column_widths) > i:
                    if width > column_widths[i]:
                        column_widths[i] = width
                else:
                    column_widths.append(width)

        for i, column_width in enumerate(column_widths):
            self.sheet.column_dimensions[get_column_letter(i + 1)].width = column_width


def read_csv(file: str):
    rows = []
    with open(file, newline='', encoding='utf-8') as csv_file:
        reader = csv.reader(csv_file)

        for row in reader:
            rows.append(row)

    return rows


def merge_report_to_xlsx(target_dir: str):
    wb = openpyxl.Workbook()
    for sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    for fn in os.listdir(target_dir):
        if not fn.endswith('.csv'):
            continue

        rows = read_csv(os.path.join(target_dir, fn))
        TableSheetGenerator(workbook=wb, title=fn.split('.')[0]).generate(headers=rows[0], rows=rows[1:])

    wb.save(os.path.join(target_dir, f'Report-{os.path.basename(target_dir)}.xlsx'))


def main():
    merge_report_to_xlsx(os.path.join('reports', '20210517_004306'))


if __name__ == '__main__':
    main()
